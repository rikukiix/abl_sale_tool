from flask import request, jsonify, send_file
from . import sale_bp
from .. import db
from ..models import Order, OrderItem, Product, Event, MasterProduct
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import io
VALID_ORDER_STATUSES = ['pending', 'completed', 'cancelled']


# --- 订单处理 API ---

# API: 获取指定展会的订单列表
# 路径: GET /sale/api/events/<int:event_id>/orders?status=pending
@sale_bp.route('/api/events/<int:event_id>/orders', methods=['GET'])
def get_orders_for_event(event_id):
    Event.query.get_or_404(event_id)
    status_filter = request.args.get('status')
    query = Order.query.filter_by(event_id=event_id)
    if status_filter in VALID_ORDER_STATUSES:
        query = query.filter_by(status=status_filter)
    orders = query.order_by(Order.timestamp.desc()).all()
    return jsonify([o.to_dict() for o in orders])

# API: 更新订单状态 (摊主操作)
# 【核心改动】API: 更新指定展会下的特定订单状态
@sale_bp.route('/api/events/<int:event_id>/orders/<int:order_id>/status', methods=['PUT'])
def update_order_status_for_event(event_id, order_id):
    data = request.get_json()
    new_status = data.get('status')

    if not new_status or new_status not in VALID_ORDER_STATUSES:
        return jsonify(error=f"Invalid status."), 400

    # 查询订单时，确保它同时匹配 order_id 和 event_id
    order = Order.query.filter_by(id=order_id, event_id=event_id).first_or_404()
    
    order.status = new_status
    db.session.commit()
    
    return jsonify(order.to_dict())
# 【新增】API: 顾客创建新订单 (替代 WebSocket)
@sale_bp.route('/api/events/<int:event_id>/orders', methods=['POST'])
def create_order(event_id):
    """
    顾客创建新订单。
    包含了健壮的库存检查，会同时考虑 'pending' 和 'completed' 的订单，
    以防止并发下单时造成的超卖问题。
    """
    data = request.get_json()
    items = data.get('items')

    if not isinstance(items, list) or not items:
        return jsonify(error="Invalid order data: Missing or empty items."), 400

    # 开启一个数据库事务，确保库存检查和订单创建的原子性
    try:
        # 在提交数据库前，先对所有请求的商品做一次完整的库存检查
        for item in items:
            product_id = item.get('product_id')
            requested_quantity = item.get('quantity')
            
            if not product_id or not requested_quantity or requested_quantity <= 0:
                
                return jsonify(error="Invalid item data in request."), 400

            product = Product.query.get(product_id)
            if not product or product.event_id != int(event_id):
                 return jsonify(error=f"Product with ID {product_id} is invalid for this event."), 400

            # 【至关重要的修改】
            # 计算所有已“承诺售出”的库存，包括已完成(completed)和待处理(pending)的。
            # 这可以防止新的订单超卖已经被其他待处理订单“预定”的库存。
            committed_quantity = db.session.query(func.sum(OrderItem.quantity))\
                .join(Order)\
                .filter(OrderItem.product_id == product_id)\
                .filter(Order.status.in_(['pending', 'completed']))\
                .scalar() or 0

            # 计算真正的可用库存
            available_stock = product.initial_stock - committed_quantity
            
            # 检查库存是否充足
            if available_stock < requested_quantity:
                
                return jsonify(
                    error=f"'{product.name}'现在库存不足.",
                    detail=f"Requested: {requested_quantity}, Available: {available_stock}"
                ), 406 # 使用 406 Not Acceptable 表示库存不足

        # --- 库存检查全部通过，现在可以安全地创建订单 ---
        
        # 为了效率，一次性获取所有涉及商品的价格
        product_ids = [i['product_id'] for i in items]
        products_in_order = Product.query.filter(Product.id.in_(product_ids)).all()
        product_prices = {p.id: p.price for p in products_in_order}

        total_amount = 0
        for item in items:
            total_amount += product_prices[item.get('product_id')] * item.get('quantity')

        new_order = Order(event_id=event_id, total_amount=round(total_amount, 2), status='pending')
        db.session.add(new_order)
        db.session.flush() # flush来获取new_order.id，以便OrderItem使用

        for item in items:
            # 创建 OrderItem 并添加到会话
            order_item = OrderItem(
                order_id=new_order.id, 
                product_id=item.get('product_id'), 
                quantity=item.get('quantity'),
                product=Product.query.get(item.get('product_id'))
            )
            db.session.add(order_item)
        
        db.session.commit() # 提交整个事务
        return jsonify(new_order.to_dict()), 201

    except Exception as e:
        db.session.rollback() # 如果发生任何错误，回滚事务
        print(f"Order creation error: {e}")
        return jsonify(error="An internal server error occurred."), 500
# --- 销售总结 API (新增功能) ---
from sqlalchemy import func
def _get_sales_summary_data(event_id):
    """
    内部辅助函数，负责查询数据库并生成销售总结的原始数据。
    【已更新】此函数现在会通过 Product JOIN MasterProduct 来获取核心商品信息。
    """
    event = Event.query.get(event_id)
    if not event:
        return None

    # 1. 【核心修改】重写数据库查询以 JOIN MasterProduct
    # 我们现在需要从 OrderItem -> Product -> MasterProduct 来获取所有信息
    # 由于 product.current_stock 是一个计算属性（通常是 @property），不能直接在 SQL 查询中选出，
    # 需要先查出 Product.id，然后在循环中用 ORM 实例获取 current_stock。
    sales_details_query = db.session.query(
        MasterProduct.product_code,  # <--- 从 MasterProduct 获取
        MasterProduct.name,          # <--- 从 MasterProduct 获取
        Product.price,               # <--- 价格仍然是展会特定的
        Product.initial_stock,       # <--- 初始库存
        func.sum(OrderItem.quantity).label('total_quantity'),
        Product.id                   # <--- 加入 Product.id 以便后续查 current_stock
    ).join(Order, OrderItem.order_id == Order.id)\
     .join(Product, OrderItem.product_id == Product.id)\
     .join(MasterProduct, Product.master_product_id == MasterProduct.id) \
     .filter(Order.event_id == event_id)\
     .filter(Order.status == 'completed')\
     .group_by(MasterProduct.product_code, MasterProduct.name, Product.price, Product.initial_stock, Product.id)\
     .order_by(MasterProduct.name)\
     .all()

    # 2. 将查询结果处理成结构化的字典列表（去重，只添加一次）
    summary_list = []
    for code, name, price, initial_stock, quantity, product_id in sales_details_query:
        product = Product.query.get(product_id)
        current_stock = product.current_stock if product else None
        summary_list.append({
            "product_code": code,
            "product_name": name,
            "unit_price": float(price),
            "initial_stock": int(initial_stock),
            "total_quantity": int(quantity),
            "total_revenue_per_item": round(float(price) * int(quantity), 2),
            "current_stock": current_stock
        })

    # 3. 计算总销售额
    total_revenue = sum(item['total_revenue_per_item'] for item in summary_list)

    # 4. 构造并返回最终的数据字典
    return {
        "event_id": event.id,
        "event_name": event.name,
        "total_revenue": round(total_revenue, 2),
        "summary": summary_list
    }


@sale_bp.route('/api/events/<int:event_id>/sales_summary', methods=['GET'])
def get_sales_summary(event_id):
    sales_data = _get_sales_summary_data(event_id)
    if sales_data is None:
        return jsonify(error="Event not found."), 404
    return jsonify(sales_data)

# 【重构】API: 下载格式精美的销售总结 Excel 文件 (现在调用辅助函数)
@sale_bp.route('/api/events/<int:event_id>/sales_summary/download', methods=['GET'])
def download_sales_summary_excel(event_id):
    """
    【已修复】
    生成并提供一份与“境界景观学会出摊标准记录”模板高度相似的 Excel 报告。
    """
    sales_data = _get_sales_summary_data(event_id)

    if sales_data is None: return "Event not found.", 404

    # --- 数据准备部分 ---
    df = pd.DataFrame(sales_data['summary'])
    if df.empty:
        df['ending_stock'] = pd.Series(dtype='object')
    else:
        # 让 ending_stock 列什么也不填（全为空）
        df['ending_stock'] = None

    column_map = {
        'product_code': '制品编号', 'product_name': '制品名', 'initial_stock': '初始数量',
        'ending_stock': '结束数量', 'unit_price': '单价', 'total_quantity': '销量',
        'total_revenue_per_item': '销售额'
    }
    df = df.rename(columns=column_map)
    df = df[list(column_map.values())]

    # --- Excel 样式定义部分保持不变 ---
    header_font = Font(name='微软雅黑', size=11, bold=True)
    body_font = Font(name='微软雅黑', size=10)
    title_font = Font(name='微软雅黑', size=16, bold=True)
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    thin_border_side = Side(style='thin', color='000000')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    center_align = Alignment(horizontal='center', vertical='center')

    # --- 创建并写入 Excel 部分 ---
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='出摊记录', index=False, header=False, startrow=2)
        ws = writer.sheets['出摊记录']

        # --- 写入和设置标题 ---
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"境界景观学会出摊标准记录 {pd.Timestamp.now().strftime('%Y年%m月')}版"
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws['G1'].value = "摊位号"
        ws['G1'].font = body_font
        ws.row_dimensions[1].height = 60

        # --- 写入和设置表头 ---
        headers = list(column_map.values())
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=i, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[2].height = 30

        for row_idx in range(3, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30
            if (row_idx % 2) == 0:
                for cell in ws[row_idx]:
                    cell.fill = light_blue_fill
            for cell in ws[row_idx]:
                cell.border = thin_border
                cell.font = body_font
                cell.alignment = center_align
            ws[f'E{row_idx}'].number_format = '¥#,##0.00'
            ws[f'G{row_idx}'].number_format = '¥#,##0.00'

        # --- 写入底部信息区域部分保持不变 ---
        bottom_start_row = ws.max_row + 2
        bottom_labels = {
            f'A{bottom_start_row}': "出席人\n与其联系方式", f'A{bottom_start_row+2}': "特殊情况\n备注",
            f'A{bottom_start_row+3}': "快递信息", f'C{bottom_start_row}': "日期",
            f'A{bottom_start_row+1}': "活动名", f'E{bottom_start_row+1}': "总销售额"
        }
        for cell_ref, label in bottom_labels.items():
            cell = ws[cell_ref]
            cell.value = label
            cell.font = body_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws[f'B{bottom_start_row+1}'].value = sales_data['event_name']
        ws.merge_cells(f'B{bottom_start_row+3}:G{bottom_start_row+3}')
        ws.merge_cells(f'B{bottom_start_row+2}:G{bottom_start_row+2}')
        ws.merge_cells(f'B{bottom_start_row+1}:D{bottom_start_row+1}')
        ws.row_dimensions[bottom_start_row+2].height = 30
        ws.row_dimensions[bottom_start_row+3].height = 30
        ws.row_dimensions[bottom_start_row+1].height = 30
        total_rev_cell = ws[f'F{bottom_start_row+1}']
        total_rev_cell.value = sales_data['total_revenue']
        total_rev_cell.number_format = '¥#,##0.00'
        total_rev_cell.font = Font(name='微软雅黑', size=12, bold=True)

        # 【核心修正】为底部区域绘制边框的逻辑
        bottom_area = f'A{bottom_start_row}:G{bottom_start_row+3}'
        for row in ws[bottom_area]:
            for cell in row:
                cell.border = thin_border

        # --- 设置固定列宽 ---
        fixed_width = 16
        for col in ws.columns:
            column = get_column_letter(col[0].column)
            ws.column_dimensions[column].width = fixed_width

    # --- 准备并发送文件部分保持不变 ---
    output_buffer.seek(0)
    filename = f"{sales_data['event_name']}_出摊记录_{pd.Timestamp.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        output_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
